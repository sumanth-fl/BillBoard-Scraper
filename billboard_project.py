#project freelance
#project 1 Billboard project

from bs4 import BeautifulSoup as soup
import urllib 
import os
import openpyxl
os.chdir('g:')
client = urllib.urlopen("http://www.billboard.com/charts/hot-100")                  
data_in_page = client.read()              
client.close()                            
soup_variable= soup(data_in_page , "html.parser")
wb = openpyxl.Workbook()
sheet = wb.active
sheet.cell(row = 1,column = 1).value = 'Song'
sheet.cell(row = 1,column = 2).value = 'Youtube link'

for num in range(0,len(soup_variable.find_all('article'))):
	song_name = soup_variable.find_all('article')[num].h2.contents
	song_list = list(str(song_name).strip('[').strip(']').strip("'"))
	song_list[0] = ''
	song_list[1] = ''
	sheet.cell(row = num+2,column = 1).value = str(''.join(song_list))
	artist_link_content = soup_variable.find_all('article')[num].a.contents
	link = "https://www.youtube.com/results?"+urllib.urlencode({'search_query':str(song_name + artist_link_content)})
	sheet.cell(row = num+2,column = 2).value = link
wb.save('billboard.xlsx')	
	
